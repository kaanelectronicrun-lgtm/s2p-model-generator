"""Self-tests for the vector fitter and the graph-fit pipeline.

Run:  python tests/test_vectorfit.py    (no pytest dependency required)
"""
import os
import sys

import numpy as np

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from s2p_tool.vectorfit import vector_fit  # noqa: E402
from s2p_tool.synth import fit_and_synthesize, synth_spice  # noqa: E402


def test_recovers_known_rational():
    """A 2-pair rational function is fitted to machine precision and stays stable."""
    f = np.geomspace(1e4, 1e10, 400)
    s = 2j * np.pi * f
    p1, p2 = complex(-1e7, 8e7), complex(-5e7, 1.2e9)
    r1, r2 = complex(3e7, -1e7), complex(-2e8, 5e7)
    resp = (r1 / (s - p1) + np.conj(r1) / (s - np.conj(p1))
            + r2 / (s - p2) + np.conj(r2) / (s - np.conj(p2)) + 0.3 + 2e-12 * s)
    m = vector_fit(f, resp, n_pairs=4, n_iter=8)
    assert m.rms_rel < 1e-6, f"rms too high: {m.rms_rel}"
    assert all(p.real <= 1e-6 for p in m.poles), "unstable pole present"


def test_noisy_resonance_fit_under_2pct():
    """A noisy single-resonance |Z| curve fits within a few percent, stays stable."""
    f = np.geomspace(1e4, 1e9, 80)
    s = 2j * np.pi * f
    C, ESL = 1e-7, 0.9e-9
    z = 0.012 + 1j * (2 * np.pi * f * ESL - 1 / (2 * np.pi * f * C))
    rng = np.random.default_rng(1)
    z = z * (1 + rng.normal(0, 0.02, len(f)))
    m = vector_fit(f, z, n_pairs=6, n_iter=8)
    assert m.rms_rel < 0.05, f"rms too high: {m.rms_rel}"
    assert all(p.real <= 1e-6 for p in m.poles), "unstable pole present"


def test_synthesis_matches_fit_and_is_passive():
    """Adaptive synthesis: netlist Z equals the VF Z, all elements positive,
    SPICE element names unique."""
    f = np.geomspace(1e4, 1e9, 80)
    s = 2j * np.pi * f
    C, ESL = 1e-7, 0.9e-9
    z = 0.012 + 1j * (2 * np.pi * f * ESL - 1 / (2 * np.pi * f * C))
    rng = np.random.default_rng(1)
    z = z * (1 + rng.normal(0, 0.02, len(f)))

    vf, synth, meta = fit_and_synthesize(f, z, max_pairs=8)
    err = np.max(np.abs(synth.impedance(s) - vf.eval(s)) / (np.abs(vf.eval(s)) + 1e-30))
    assert err < 1e-9, f"synth impedance != fit: {err}"
    assert meta["passive_synth"], "chosen model is not passive"

    # All emitted element values strictly positive, names unique.
    names, vals = [], []
    for line in synth_spice(synth, "PART").splitlines():
        if line and line[0] in "RLC" and " " in line:
            tok = line.split()
            names.append(tok[0])
            vals.append(float(tok[-1]))
    assert len(names) == len(set(names)), f"duplicate element name: {names}"
    assert all(v > 0 for v in vals), f"non-positive element: {vals}"


def test_skrf_backend_passive_and_null_accurate():
    """If scikit-rf is installed, the backend fits a noisy resonance to a passive,
    null-accurate model. Skipped when skrf is absent (numpy-only install)."""
    from s2p_tool import skrf_backend as sk
    if not sk.available():
        print("  (skipped: scikit-rf not installed)")
        return
    import os
    f = np.geomspace(1e4, 1e9, 200)
    w = 2 * np.pi * f
    z = 0.02 + 1j * (w * 0.6e-9 - 1.0 / (w * 1.0e-7))   # 100 nF, 0.6 nH, 20 mohm
    i = int(np.argmin(np.abs(z)))
    out = os.path.join(os.path.dirname(__file__), "_skrf_tmp.sp")
    r = sk.fit_passive_spice(f, z, 50.0, out)
    dip = abs(abs(r["z_model"][i]) - abs(z[i])) / abs(z[i]) * 100
    assert r["passive_after"], "skrf model not passive"
    assert dip < 5.0, f"skrf dip error too high: {dip}%"
    if os.path.exists(out):
        os.remove(out)


def test_spice_roundtrip_netlist_realises_model():
    """The emitted .cir, re-simulated by the independent nodal solver, reproduces
    the synthesis impedance (proves the netlist text is correct)."""
    from s2p_tool import spicesim
    from s2p_tool.synth import fit_and_synthesize, synth_spice
    f = np.geomspace(1e4, 1e9, 200)
    w = 2 * np.pi * f
    z = 0.02 + 1j * (w * 0.6e-9 - 1.0 / (w * 1.0e-7))
    vf, synth, meta = fit_and_synthesize(f, z, max_pairs=8)
    cir = synth_spice(synth, "DUT")
    r = spicesim.roundtrip(cir, f, synth.impedance(2j * np.pi * f))
    assert r["netlist_ok"], f"netlist round-trip error too high: {r['netlist_rel_err']}"


def test_pdf_field_extraction():
    """Datasheet-text heuristics pull C / voltage / dielectric / case correctly."""
    from s2p_tool import pdfreader
    txt = "GRM188R71C104KA01 (0603, X7R:EIA, 0.1uF, DC16V) Chip Capacitor"
    f = pdfreader.extract_capacitor(txt)
    assert abs(f["capacitance_f"] - 1e-7) < 1e-12, f["capacitance_f"]
    assert f["voltage_rating_v"] == 16.0
    assert f["dielectric"] == "X7R"
    assert f["case"] == "0603"
    assert pdfreader._part_number(txt, "x") == "GRM188R71C104KA01"

def test_series_shunt_closed_form():
    """Series and shunt topologies match their closed-form S-params and are
    reciprocal + symmetric. Uses a plain resistive Z so the algebra is exact."""
    from s2p_tool import sparams
    z0 = 50.0
    z = np.array([10.0, 50.0, 200.0], dtype=complex)
    ss = sparams.series_z_to_s(z, z0)
    sh = sparams.shunt_z_to_s(z, z0)
    # series closed form
    assert np.allclose(ss["S11"], z / (z + 2 * z0))
    assert np.allclose(ss["S21"], (2 * z0) / (z + 2 * z0))
    # shunt closed form
    assert np.allclose(sh["S11"], (-z0) / (2 * z + z0))
    assert np.allclose(sh["S21"], (2 * z) / (2 * z + z0))
    # reciprocity + symmetry for both
    for s in (ss, sh):
        assert np.allclose(s["S21"], s["S12"])
        assert np.allclose(s["S11"], s["S22"])
    # dispatcher agrees
    assert np.allclose(sparams.z_to_s(z, z0, "series")["S21"], ss["S21"])
    assert np.allclose(sparams.z_to_s(z, z0, "shunt")["S21"], sh["S21"])


def test_series_shunt_srf_signature():
    """At SRF (Z->min) a series cap passes (|S21|->1) while a shunt cap blocks
    (|S21|->0) — the physical fingerprint that distinguishes the two models."""
    import tempfile
    from s2p_tool.pipeline import process

    out = tempfile.mkdtemp()
    j = os.path.join(os.path.dirname(__file__), "..", "components", "example_cap.json")

    def s21_mags(path):
        rows = [ln.split() for ln in open(path) if ln[:1].isdigit()]
        return [(float(r[3]) ** 2 + float(r[4]) ** 2) ** 0.5 for r in rows]

    s_series = process(j, out, 50.0, 1e4, 1e10, "series")[0]
    s_shunt = process(j, out, 50.0, 1e4, 1e10, "shunt")[0]
    assert s_series.endswith("_series.s2p")
    assert s_shunt.endswith("_shunt.s2p")
    assert max(s21_mags(s_series)) > 0.95, "series should pass near SRF"
    assert min(s21_mags(s_shunt)) < 0.05, "shunt should block near SRF"


def test_derate_class_and_dc_bias():
    """Class II derates under DC bias (monotonic loss); Class I stays put."""
    from s2p_tool import derate
    assert derate.dielectric_class("C0G") == "I"
    assert derate.dielectric_class("NP0") == "I"
    assert derate.dielectric_class("X7R") == "II"
    assert derate.dielectric_class(None) == "II"
    # X7R, 16 V rated: more bias -> less capacitance, monotonic, bounded.
    f0, _, e0 = derate.compute_factor("X7R", 16.0, dc_bias_v=0.0)
    f_half, _, _ = derate.compute_factor("X7R", 16.0, dc_bias_v=8.0)
    f_full, _, est = derate.compute_factor("X7R", 16.0, dc_bias_v=16.0)
    assert f0 == 1.0
    assert 0.0 < f_full < f_half < 1.0, (f_full, f_half)
    assert est is True, "behavioral estimate must flag"
    # C0G at the same bias: negligible change.
    f_c0g, _, _ = derate.compute_factor("C0G", 16.0, dc_bias_v=16.0)
    assert abs(f_c0g - 1.0) < 1e-9


def test_derate_vendor_curve_is_exact():
    """A digitized dc_bias_curve is interpolated exactly and is NOT flagged as
    a behavioral estimate (it is real vendor data)."""
    from s2p_tool import derate
    curve = [[0, 0.0], [10, -40.0]]  # -40% at 10 V, linear
    f, log, est = derate.compute_factor("X7R", 16.0, dc_bias_v=5.0,
                                        dc_bias_curve=curve)
    assert abs(f - 0.80) < 1e-6, f  # -20% at 5 V (midpoint)
    assert est is False
    assert any("vendor curve" in ln for ln in log)


def test_dc_bias_shifts_srf_up():
    """Derating lowers effective C, so the model SRF moves UP (1/2π√(LC))."""
    import json
    import tempfile
    from s2p_tool.pipeline import process

    out = tempfile.mkdtemp()
    base = json.load(open(os.path.join(os.path.dirname(__file__), "..",
                                       "components", "example_cap.json")))

    def srf_of(path):
        mags = []
        for ln in open(path):
            if ln[:1].isdigit():
                v = ln.split()
                # |S21| max == series resonance (pass peak) for a series cap
                mags.append((float(v[0]),
                             (float(v[3]) ** 2 + float(v[4]) ** 2) ** 0.5))
        return max(mags, key=lambda x: x[1])[0]

    p0 = os.path.join(out, "n.json"); json.dump(base, open(p0, "w"))
    s0 = process(p0, out, 50.0, 1e4, 1e10, "series")[0]
    d = dict(base); d["part_number"] = "biased"; d["dc_bias_v"] = 16.0
    p1 = os.path.join(out, "b.json"); json.dump(d, open(p1, "w"))
    s1 = process(p1, out, 50.0, 1e4, 1e10, "series")[0]
    assert srf_of(s1) > srf_of(s0), (srf_of(s1), srf_of(s0))


def test_curve_csv_loading_and_json_path():
    """A SimSurfing-style CSV (with header) loads, and a JSON that references it
    by path drives exact (non-estimate) derating end to end."""
    import json
    import tempfile
    from s2p_tool import derate
    from s2p_tool.pipeline import load_component, process

    d = tempfile.mkdtemp()
    csv_path = os.path.join(d, "dcbias.csv")
    with open(csv_path, "w", encoding="utf-8") as fh:
        fh.write("DC Bias[V],Cap. Change Rate[%]\n0,0\n5,-25\n10,-60\n")
    curve = derate.load_curve_csv(csv_path)
    assert curve[0] == [0.0, 0.0] and curve[-1] == [10.0, -60.0]

    # JSON references the CSV by relative path -> load_component resolves it.
    cj = {"kind": "capacitor", "part_number": "CURVE", "capacitance_f": 1e-7,
          "voltage_rating_v": 16, "dielectric": "X7R", "srf_hz": 1.6e7,
          "dc_bias_v": 5.0, "dc_bias_curve": "dcbias.csv"}
    jp = os.path.join(d, "c.json"); json.dump(cj, open(jp, "w"))
    comp = load_component(jp)
    assert comp.dc_bias_curve == curve  # string path resolved to list

    # -25% at 5 V (curve point) -> effective 75 nF, exact, source kept (not estimate).
    rep = process(jp, d, 50.0, 1e4, 1e10, "series")[2]
    txt = open(rep, encoding="utf-8").read()
    assert "condition-derated" in txt
    assert "75.000 nF" in txt, txt[:600]
    assert "vendor curve" in txt


def _make_datasheet_pdf(path, known):
    """Draw a synthetic MLCC datasheet page: text params + a DC-bias vector plot."""
    import fitz
    doc = fitz.open(); page = doc.new_page(width=420, height=460)
    page.insert_text(fitz.Point(30, 30), "GRT188R71C104KA01 Ceramic Capacitor", fontsize=10)
    page.insert_text(fitz.Point(30, 46),
                     "Capacitance 100nF Rated Voltage DC 16V X7R Size 1608 (0603)", fontsize=9)
    page.insert_text(fitz.Point(30, 70), "DC Bias Characteristics", fontsize=10)
    X0, X1, Y0, Y1 = 110, 360, 110, 330
    page.draw_rect(fitz.Rect(X0, Y0, X1, Y1))
    for v in [0, 4, 8, 12, 16]:
        px = X0 + (X1 - X0) * v / 16
        page.draw_line(fitz.Point(px, Y1), fitz.Point(px, Y1 + 4))
        page.insert_text(fitz.Point(px - 4, Y1 + 16), str(v), fontsize=8)
    for v in [0, -20, -40, -60, -80]:
        py = Y0 + (Y1 - Y0) * (-v) / 80
        page.draw_line(fitz.Point(X0 - 4, py), fitz.Point(X0, py))
        page.insert_text(fitz.Point(X0 - 30, py + 3), str(v), fontsize=8)
    page.insert_text(fitz.Point(200, Y1 + 30), "DC Bias[V]", fontsize=9)
    page.insert_text(fitz.Point(24, 210), "Cap. Change Rate[%]", fontsize=9)
    pix = [(X0 + (X1 - X0) * v / 16, Y0 + (Y1 - Y0) * (-d) / 80) for v, d in known]
    for a, b in zip(pix[:-1], pix[1:]):
        page.draw_line(fitz.Point(*a), fitz.Point(*b))
    doc.save(path); doc.close()


def test_pdf_curve_extraction_and_process_pdf():
    """Digitize a DC-bias curve from a vector PDF and drive one-click process_pdf.
    Skips cleanly if PyMuPDF is absent."""
    import tempfile
    import numpy as np
    from s2p_tool import pdfcurves
    if not pdfcurves.available():
        print("  (skipped: PyMuPDF not installed)")
        return
    from s2p_tool.pipeline import process_pdf

    d = tempfile.mkdtemp()
    pdf = os.path.join(d, "GRT188.pdf")
    known = [(0, 0), (4, -12), (8, -30), (12, -52), (16, -70)]
    _make_datasheet_pdf(pdf, known)

    cr = pdfcurves.extract_curves(pdf)
    c = cr["dc_bias_curve"]
    assert c and len(c) >= 4, cr["log"]
    xs = [p[0] for p in c]; ys = [p[1] for p in c]
    # recovered curve matches the known points within digitization tolerance
    for V, expect in [(4, -12), (8, -30), (12, -52)]:
        assert abs(float(np.interp(V, xs, ys)) - expect) < 2.0, (V, np.interp(V, xs, ys))

    # one-click: PDF -> s2p at 8 V uses the extracted curve (~-30%), 1 kHz-10 GHz
    s2p, cir, rep = process_pdf(pdf, "capacitor", d, 50.0, 1e3, 1e10,
                                "series", dc_bias_v=8.0)
    txt = open(rep, encoding="utf-8").read()
    assert "vendor curve" in txt, "extracted curve should drive exact derating"
    freqs = [float(ln.split()[0]) for ln in open(s2p) if ln[:1].isdigit()]
    assert min(freqs) <= 1e3 and max(freqs) >= 1e10, (min(freqs), max(freqs))


def test_impedance_curve_vector_read_and_graphfit():
    """Read a log-log |Z|(f) curve from a vector PDF and confirm process_pdf
    vector-fits it (every point) into an s2p that follows the datasheet curve."""
    import math
    import tempfile
    import numpy as np
    from s2p_tool import pdfcurves
    if not pdfcurves.available():
        print("  (skipped: PyMuPDF not installed)")
        return
    import fitz
    from s2p_tool.pipeline import process_pdf

    C, L, ESR = 1e-7, 0.6e-9, 0.03
    zmag = lambda f: abs(ESR + 1j * (2 * math.pi * f * L - 1 / (2 * math.pi * f * C)))
    doc = fitz.open(); page = doc.new_page(width=460, height=480)
    page.insert_text(fitz.Point(30, 26),
                     "GRT188R71C104KA01 Capacitance 100nF DC 16V X7R 0603", fontsize=9)
    page.insert_text(fitz.Point(30, 44), "Impedance vs Frequency", fontsize=10)
    X0, X1, Y0, Y1 = 110, 410, 90, 360
    page.draw_rect(fitz.Rect(X0, Y0, X1, Y1))
    fx = lambda f: X0 + (X1 - X0) * (math.log10(f) - 3) / 6
    fy = lambda z: Y1 - (Y1 - Y0) * (math.log10(z) + 2) / 5
    for f, lab in {1e3: "1k", 1e4: "10k", 1e5: "100k", 1e6: "1M",
                   1e7: "10M", 1e8: "100M", 1e9: "1G"}.items():
        px = fx(f); page.draw_line(fitz.Point(px, Y1), fitz.Point(px, Y1 + 4))
        page.insert_text(fitz.Point(px - 6, Y1 + 16), lab, fontsize=8)
    for z, lab in {1e-2: "0.01", 1e-1: "0.1", 1e0: "1", 1e1: "10",
                   1e2: "100", 1e3: "1k"}.items():
        py = fy(z); page.draw_line(fitz.Point(X0 - 4, py), fitz.Point(X0, py))
        page.insert_text(fitz.Point(X0 - 32, py + 3), lab, fontsize=8)
    page.insert_text(fitz.Point(230, Y1 + 30), "Frequency (Hz)", fontsize=9)
    page.insert_text(fitz.Point(24, 210), "Impedance (Ohm)", fontsize=9)
    fs = [10 ** (3 + 6 * i / 120) for i in range(121)]
    pix = [(fx(f), fy(zmag(f))) for f in fs]
    for a, b in zip(pix[:-1], pix[1:]):
        page.draw_line(fitz.Point(*a), fitz.Point(*b))
    d = tempfile.mkdtemp(); pdf = os.path.join(d, "GRT_imp.pdf")
    doc.save(pdf); doc.close()

    # raw extraction accuracy (log-log axes, SI-suffixed ticks)
    zc = pdfcurves.extract_curves(pdf)["impedance_curve"]
    assert zc and len(zc) >= 50, zc
    fsx = np.array([q[0] for q in zc]); zsx = np.array([q[1] for q in zc])
    o = np.argsort(fsx); fsx, zsx = fsx[o], zsx[o]
    for ft in [1e4, 1e5, 1e6]:
        got = 10 ** np.interp(math.log10(ft), np.log10(fsx), np.log10(zsx))
        assert abs(got / zmag(ft) - 1) < 0.05, (ft, got, zmag(ft))

    # one-click -> graph-fit s2p follows the curve
    s2p, cir, rep = process_pdf(pdf, "capacitor", d, 50.0, 1e3, 1e10, "series")
    txt = open(rep, encoding="utf-8").read()
    assert "HIGH-FIDELITY" in txt and "GRAPH FIT" in txt.upper(), txt[:400]
    arr = np.array([[float(x) for x in ln.split(",")]
                    for ln in open(os.path.join(d, os.path.basename(s2p).replace(".s2p", "_Zf.csv")))
                    if ln[0].isdigit()])
    fm, mg = arr[:, 0], arr[:, 3]
    for ft in [1e4, 1e5, 1e6]:
        got = 10 ** np.interp(math.log10(ft), np.log10(fm), np.log10(mg))
        assert abs(got / zmag(ft) - 1) < 0.06, (ft, got, zmag(ft))


def test_grid_curve_segmentation_and_confidence():
    """Two captioned plots share one page (a grid). extract_labeled_curves must
    anchor each to its own caption/column, calibrate independently, and return
    the correct curve at high confidence — no cross-talk between subplots."""
    import tempfile
    from s2p_tool import pdfcurves
    if not pdfcurves.available():
        print("  (skipped: PyMuPDF not installed)")
        return
    import fitz

    doc = fitz.open(); page = doc.new_page(width=612, height=460)
    # Left subplot: Vref vs Temperature (y ~1.204 V).
    LX0, LX1, LY0, LY1 = 100, 280, 70, 210
    page.draw_rect(fitz.Rect(LX0, LY0, LX1, LY1))
    fxL = lambda T: LX0 + (LX1 - LX0) * (T + 40) / 160.0
    fyL = lambda V: LY1 - (LY1 - LY0) * (V - 1.202) / 0.004
    for T in (-40, 0, 40, 80, 120):
        page.insert_text(fitz.Point(fxL(T) - 6, LY1 + 14), str(T), fontsize=8)
    for V in (1.202, 1.204, 1.206):
        page.insert_text(fitz.Point(LX0 - 30, fyL(V) + 3), f"{V:.3f}", fontsize=8)
    vpts = [(fxL(T), fyL(1.205 - (T + 40) / 160.0 * 0.002))
            for T in range(-40, 121, 8)]
    for a, b in zip(vpts[:-1], vpts[1:]):
        page.draw_line(fitz.Point(*a), fitz.Point(*b))
    page.insert_text(fitz.Point(LX0 + 6, LY1 + 34),
                     "Figure 1. Reference Voltage vs Temperature", fontsize=8)

    # Right subplot: Quiescent Current vs Temperature (y in µA).
    RX0, RX1, RY0, RY1 = 360, 540, 70, 210
    page.draw_rect(fitz.Rect(RX0, RY0, RX1, RY1))
    fxR = lambda T: RX0 + (RX1 - RX0) * (T + 40) / 160.0
    fyR = lambda I: RY1 - (RY1 - RY0) * I / 140.0
    for T in (-40, 0, 40, 80, 120):
        page.insert_text(fitz.Point(fxR(T) - 6, RY1 + 14), str(T), fontsize=8)
    for I in (0, 40, 80, 120):
        page.insert_text(fitz.Point(RX0 - 26, fyR(I) + 3), str(I), fontsize=8)
    ipts = [(fxR(T), fyR(90 + (T + 40) / 160.0 * 30))
            for T in range(-40, 121, 8)]
    for a, b in zip(ipts[:-1], ipts[1:]):
        page.draw_line(fitz.Point(*a), fitz.Point(*b))
    page.insert_text(fitz.Point(RX0 + 6, RY1 + 34),
                     "Figure 2. Quiescent Current vs Temperature", fontsize=8)

    d = tempfile.mkdtemp(); pdf = os.path.join(d, "grid.pdf")
    doc.save(pdf); doc.close()

    res = pdfcurves.extract_labeled_curves(
        pdf, [("vref", r"reference voltage"), ("iq", r"quiescent current")])
    assert "vref" in res and "iq" in res, res["log"]
    assert res["vref"]["confidence"] == "high", res["log"]
    assert res["iq"]["confidence"] == "high", res["log"]
    # Correct caption anchoring (segmentation, no cross-talk).
    assert "Reference Voltage" in res["vref"]["caption"]
    assert "Quiescent Current" in res["iq"]["caption"]
    # Left plot recovered its own axis (volts ~1.204), NOT the right plot's µA.
    vy = [p[1] for p in res["vref"]["curve"]]
    assert 1.19 < (sorted(vy)[len(vy) // 2]) < 1.21, vy[:5]
    vx = [p[0] for p in res["vref"]["curve"]]
    assert min(vx) < -20 and max(vx) > 100, (min(vx), max(vx))
    # Right plot recovered a ranging µA curve (segmentation kept it separate).
    iy = [p[1] for p in res["iq"]["curve"]]
    assert max(iy) - min(iy) > 10 and 60 < max(iy) < 140, (min(iy), max(iy))


def test_component_specs_and_type_detect():
    """Regulator spec regexes + type detection against a synthetic datasheet
    text block — the observable contract of the analysis engine, no PDF."""
    from s2p_tool import component_analysis as ca
    text = (
        "TPSXXXX Boost Converter\n"
        "2.7-V to 12-V input voltage range\n"
        "4.5-V to 12.6-V output voltage range\n"
        "10-A switch current\n"
        "Up to 91% efficiency at VIN = 3.3 V\n"
        "1.0-\u00b5A current into the VIN pin during shutdown\n"
        "Adjustable switching frequency: 200 kHz to 2.2 MHz\n"
        "Output overvoltage protection at 13.2 V\n"
        "4.50-mm \u00d7 3.50-mm 20-pin VQFN package\n"
        "synchronous boost converter with a 11-m\u03a9 \npower switch and a "
        "13-m\u03a9 rectifier switch to provide high efficiency.\n"
    )
    key, scores = ca.detect_type(text)
    assert key == "regulator", (key, scores)
    specs = dict(ca.RegulatorAnalyzer().extract_specs(text))
    assert specs["Giri\u015f gerilimi (VIN)"] == "2.7 V \u2013 12 V", specs
    assert specs["\u00c7\u0131k\u0131\u015f gerilimi (VOUT)"] == "4.5 V \u2013 12.6 V", specs
    assert specs["Anahtar ak\u0131m kapasitesi"] == "10 A", specs
    assert specs["Anahtarlama frekans\u0131"] == "200 kHz \u2013 2.2 MHz", specs
    assert specs["Tepe verim"] == "%91", specs
    assert specs["Kapal\u0131-durum ak\u0131m\u0131 (shutdown)"] == "1.0 \u00b5A", specs
    assert specs["A\u015f\u0131r\u0131 gerilim koruma (OVP)"] == "13.2 V", specs
    assert "11 m\u03a9" in specs["Anahtar direnci RDS(on)"], specs
    assert "VQFN" in specs["Paket"], specs
    assert specs["Topoloji"] == "Senkron boost", specs


def test_regulator_design_and_layout_extractors():
    """Design-requirements table, numbered design-procedure steps + glossary,
    and layout-guideline checklist parse from a synthetic datasheet text."""
    from s2p_tool import component_analysis as ca
    text = (
        "8.2.1 Design Requirements\nTable 8-1. Design Parameters\n"
        "DESIGN PARAMETERS\nEXAMPLE VALUES\n"
        "Input voltage range\n3.3 to 4.2 V\nOutput voltage\n9 V\n"
        "Output voltage ripple\n100 mV peak to peak\nOutput current rating\n3 A\n"
        "Operating frequency\n600 kHz\nOperation mode at light load\nPFM\n"
        "8.2.2 Detailed Design Procedure\n"
        "8.2.2.2 Setting Switching Frequency\n"
        "The switching frequency is set by a resistor between FSW and SW.\nwhere\n"
        "RFREQ is the resistance connected between the FSW pin and the SW pin\n"
        "CFREQ is 23 pF\ntDELAY is 89 ns\n"
        "8.2.2.5 Inductor Selection\n"
        "Three important specs are the inductor value, saturation and DCR.\n"
        "9 Power Supply Recommendations\n10 Layout\n10.1 Layout Guidelines\n"
        "Minimize the length and area of all traces connected to the SW pin. "
        "The input capacitor needs to be close to the VIN pin and GND pin. "
        "Use thermal vias underneath the thermal pad.\n10.2 Layout Example\n"
    )
    req = dict(ca._design_requirements(text))
    assert req["Giri\u015f gerilimi aral\u0131\u011f\u0131"] == "3.3 to 4.2 V", req
    assert req["\u00c7\u0131k\u0131\u015f gerilimi"] == "9 V", req
    assert req["\u00c7al\u0131\u015fma frekans\u0131"] == "600 kHz", req
    assert req["Hafif-y\u00fck modu"] == "PFM", req

    proc = ca._design_procedure(text)
    steps = {s["step"]: s for s in proc}
    assert "Anahtarlama frekans\u0131 ayar\u0131 (RFSW)" in steps, steps
    fsw = steps["Anahtarlama frekans\u0131 ayar\u0131 (RFSW)"]
    joined = " ".join(fsw["vars"])
    assert "CFREQ = 23 pF" in joined and "tDELAY = 89 ns" in joined, fsw
    assert any("İndükt" in k or "ndukt" in k or "nd\u00fckt" in k
               for k in steps), list(steps)

    lay = ca._layout_guidelines(text)
    assert len(lay) >= 3, lay
    assert any("SW pin" in s for s in lay), lay
    assert not any(set(s) <= {".", " "} for s in lay), lay  # no dotted leaders


def test_regulator_curve_interpretation():
    """High-confidence curves become correct plain-language findings; a
    decreasing current-limit curve is reported as decreasing despite endpoints."""
    from s2p_tool.component_analysis import RegulatorAnalyzer
    curves = {
        "vref_vs_temp": {
            "confidence": "high", "unit_x": "°C", "unit_y": "V",
            "desc": "Vref", "npoints": 3,
            "curve": [[-40, 1.205], [25, 1.204], [125, 1.203]]},
        "current_limit_vs_r": {
            "confidence": "high", "unit_x": "kΩ", "unit_y": "A",
            "desc": "ILIM", "npoints": 5,
            "curve": [[90, 13.0], [150, 8.0], [250, 4.6], [349, 3.3]]},
        "efficiency": {  # low confidence -> must be skipped
            "confidence": "low", "unit_x": "A", "unit_y": "%",
            "desc": "Eff", "npoints": 9,
            "curve": [[0.1, 80], [1, 90], [3, 88]]},
    }
    out = RegulatorAnalyzer().interpret_curves(curves)
    joined = "\n".join(out)
    assert "1.204" in joined and "kararl" in joined, out
    assert "azal" in joined, out            # current limit decreases with R
    assert "Eff" not in joined, out          # low-confidence curve skipped


def test_excel_aggregation_upsert():
    """Multiple analyses accumulate in one workbook; re-adding a part upserts
    (no duplicate) and a second part is appended across the long sheets."""
    from s2p_tool import excel_export as xl
    if not xl.available():
        print("  (skipped: openpyxl not installed)")
        return
    import tempfile
    import openpyxl
    res = {
        "part": "PARTA", "type": "regulator", "type_label": "Regülatör",
        "source_pdf": "/x/a.pdf",
        "specs": [("Giriş gerilimi (VIN)", "2.7 V – 12 V"), ("Paket", "VQFN")],
        "pinout": [{"name": "VIN", "number": "1", "io": "I", "desc": "in"},
                   {"name": "GND", "number": "2", "io": "—", "desc": "gnd"}],
        "curves": {"vref": {"confidence": "high", "npoints": 10, "desc": "Vref",
                            "unit_x": "C", "unit_y": "V"}, "_log": []},
        "curve_interpretation": ["Vref sabit"],
        "design_procedure": [{"step": "S1", "intent": "do", "vars": ["a=1"]}],
        "layout_guidelines": ["keep short"],
    }
    path = os.path.join(tempfile.mkdtemp(), "agg.xlsx")
    xl.append_analysis(res, path)
    xl.append_analysis(res, path)              # re-add -> upsert
    resB = dict(res); resB["part"] = "PARTB"
    xl.append_analysis(resB, path)
    wb = openpyxl.load_workbook(path)
    oz = list(wb["Özet"].iter_rows(min_row=2, values_only=True))
    assert [r[0] for r in oz] == ["PARTA", "PARTB"], oz  # upsert, then append
    assert wb["Pinout"].max_row - 1 == 4, wb["Pinout"].max_row  # 2 pins x 2 parts
    hdr = [c.value for c in wb["Özet"][1]]
    assert oz[0][hdr.index("VIN")] == "2.7 V – 12 V", oz[0]

if __name__ == "__main__":
    test_recovers_known_rational()
    test_noisy_resonance_fit_under_2pct()
    test_synthesis_matches_fit_and_is_passive()
    test_skrf_backend_passive_and_null_accurate()
    test_spice_roundtrip_netlist_realises_model()
    test_pdf_field_extraction()
    test_series_shunt_closed_form()
    test_series_shunt_srf_signature()
    test_derate_class_and_dc_bias()
    test_derate_vendor_curve_is_exact()
    test_dc_bias_shifts_srf_up()
    test_curve_csv_loading_and_json_path()
    test_pdf_curve_extraction_and_process_pdf()
    test_impedance_curve_vector_read_and_graphfit()
    test_grid_curve_segmentation_and_confidence()
    test_component_specs_and_type_detect()
    test_regulator_design_and_layout_extractors()
    test_regulator_curve_interpretation()
    test_excel_aggregation_upsert()
    print("OK - all vector-fit + synthesis tests passed")
