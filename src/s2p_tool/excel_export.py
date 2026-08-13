"""Aggregate component analyses into ONE Excel workbook — the 'büyük projeler
için topla' step. Each `analyze_pdf` result is upserted by part number across a
set of sheets, so re-analyzing a part refreshes its rows instead of duplicating
them and a project's whole component set accumulates in a single file.

Sheets
------
- Özet          : one wide row per part (curated key specs + counts).
- Speclar       : long (Parça, Parametre, Değer) — schema-free, any type.
- Pinout        : long pin table across all parts.
- Egriler       : extracted characteristic curves (confidence, range).
- GrafikYorumu  : plain-language curve findings.
- Hesaplar      : design-procedure steps + variable glossary.
- Layout        : datasheet layout-guideline checklist.
"""
from __future__ import annotations

import os
from typing import Dict, List, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    _HAVE = True
except Exception:  # pragma: no cover
    _HAVE = False


def available() -> bool:
    return _HAVE


# Curated wide-summary columns: (Özet header, spec label as stored in result).
_SUMMARY_SPECS: Tuple[Tuple[str, str], ...] = (
    ("VIN", "Giriş gerilimi (VIN)"),
    ("VOUT", "Çıkış gerilimi (VOUT)"),
    ("I_switch", "Anahtar akım kapasitesi"),
    ("fSW", "Anahtarlama frekansı"),
    ("Verim", "Tepe verim"),
    ("Iq_shutdown", "Kapalı-durum akımı (shutdown)"),
    ("OVP", "Aşırı gerilim koruma (OVP)"),
    ("RDS(on)", "Anahtar direnci RDS(on)"),
    ("Paket", "Paket"),
    ("Topoloji", "Topoloji"),
)

_SUMMARY_HEADER = (["Parça", "Tür"] + [h for h, _ in _SUMMARY_SPECS]
                   + ["#pin", "#eğri(high)", "Kaynak"])


def _sheet(wb, name: str, header: List[str]):
    """Get or create a sheet, ensuring the header row exists."""
    if name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row >= 1 and any(c.value for c in ws[1]):
            return ws
    else:
        ws = wb.create_sheet(name)
    ws.delete_rows(1, ws.max_row or 1)
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    return ws


def _upsert(ws, part: str, header: List[str], rows: List[List]):
    """Replace every existing row for `part` (col A) with `rows`, keep the rest."""
    kept = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and str(r[0]) != str(part):
            kept.append(list(r))
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for r in kept + rows:
        ws.append(r)


def _autofit(ws, cap: int = 70):
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, 1):
            widths[i] = max(widths.get(i, 0), len(str(v)) if v is not None else 0)
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(cap, max(8, w + 2))


def append_analysis(result: Dict, xlsx_path: str) -> str:
    """Upsert one analysis result into the aggregate workbook at `xlsx_path`,
    creating it if absent. Returns the path. Keyed by result['part']."""
    if not _HAVE:
        raise RuntimeError("openpyxl gerekli — kurulu değil.")
    part = str(result.get("part", "component"))
    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)) or ".", exist_ok=True)

    if os.path.isfile(xlsx_path):
        wb = load_workbook(xlsx_path)
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row <= 1:
            del wb["Sheet"]
    else:
        wb = Workbook()
        del wb[wb.sheetnames[0]]  # drop default empty sheet

    specs = dict(result.get("specs", []))
    curves = result.get("curves", {})
    data_curves = [(k, c) for k, c in curves.items()
                   if k != "_log" and isinstance(c, dict) and "npoints" in c]
    n_high = sum(1 for _, c in data_curves if c.get("confidence") == "high")
    src = os.path.basename(result.get("source_pdf", ""))

    # --- Özet (wide, one row) ---
    ws = _sheet(wb, "Özet", _SUMMARY_HEADER)
    row = [part, result.get("type_label", result.get("type", ""))]
    row += [specs.get(lbl, "") for _, lbl in _SUMMARY_SPECS]
    row += [len(result.get("pinout", [])), n_high, src]
    _upsert(ws, part, _SUMMARY_HEADER, [row])

    # --- Speclar (long, schema-free) ---
    hdr = ["Parça", "Parametre", "Değer"]
    ws = _sheet(wb, "Speclar", hdr)
    _upsert(ws, part, hdr, [[part, lbl, val] for lbl, val in result.get("specs", [])])

    # --- Pinout ---
    hdr = ["Parça", "Pin", "No", "I/O", "Açıklama"]
    ws = _sheet(wb, "Pinout", hdr)
    _upsert(ws, part, hdr,
            [[part, p["name"], p["number"], p["io"], p["desc"]]
             for p in result.get("pinout", [])])

    # --- Egriler (one row per trace; single-trace plots -> one "tek" row) ---
    hdr = ["Parça", "Eğri", "İz", "Güven", "Nokta", "X", "Y"]
    ws = _sheet(wb, "Egriler", hdr)
    egri_rows = []
    for _, c in data_curves:
        if c.get("traces"):
            for tr in c["traces"]:
                egri_rows.append([part, c["desc"], tr["label"], c["confidence"],
                                  tr["npoints"], c["unit_x"], c["unit_y"]])
        else:
            egri_rows.append([part, c["desc"], "tek", c["confidence"],
                              c["npoints"], c["unit_x"], c["unit_y"]])
    _upsert(ws, part, hdr, egri_rows)

    # --- Grafik Yorumu ---
    hdr = ["Parça", "Yorum"]
    ws = _sheet(wb, "GrafikYorumu", hdr)
    _upsert(ws, part, hdr,
            [[part, s] for s in result.get("curve_interpretation", [])])

    # --- Hesaplar (design procedure) ---
    hdr = ["Parça", "Adım", "Amaç", "Değişkenler / sabitler"]
    ws = _sheet(wb, "Hesaplar", hdr)
    _upsert(ws, part, hdr,
            [[part, s["step"], s.get("intent", ""), "\n".join(s.get("vars", []))]
             for s in result.get("design_procedure", [])])

    # --- Layout ---
    hdr = ["Parça", "Layout önerisi"]
    ws = _sheet(wb, "Layout", hdr)
    _upsert(ws, part, hdr,
            [[part, s] for s in result.get("layout_guidelines", [])])

    for ws in wb.worksheets:
        _autofit(ws)
    wb.save(xlsx_path)
    return xlsx_path
